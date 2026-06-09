"""
DBF Segregator — High-performance CLI tool
Splits DBF files by organisation using a master XLSX PAN->ORG mapping.

Usage:
    py segregate.py --master master.xlsx --dbf file1.dbf file2.dbf file3.dbf
    py segregate.py --master master.xlsx --dbf file1.dbf --output ./output
    py segregate.py --master master.xlsx --dbf file1.dbf --sheet May22
    py segregate.py --help

Requirements:
    pip install -r requirements.txt
"""

import argparse
import sys
import time
import struct
import logging
from pathlib import Path
from datetime import date
from concurrent.futures import ProcessPoolExecutor, as_completed

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("segregator")


# ════════════════════════════════════════════════════════════════════════════
#  CONFIG — Possible fallback field name pairs (folio, product)
#  The script auto-detects which pair exists in each DBF/CSV header.
#  Add more pairs here if needed.
# ════════════════════════════════════════════════════════════════════════════
FOLIO_PRODUCT_CANDIDATES = [
    ("FOLIO_NO",      "PRODCODE"),        # DBF1, DBF2
    ("FOLIOCHK",      "PRODUCT"),         # DBF3
    ("Folio Number",  "Product code"),    # CSV1, CSV2
    ("Folio",         "Product code"),    # CSV3
]

# PAN column name candidates for CSV files (case-insensitive match)
CSV_PAN_CANDIDATES = ["PAN1", "PAN Number", "PAN_NO", "PAN"]

def get_folio_product_fields(fields: list):
    """
    Auto-detect folio+product field pair by checking actual DBF field names.
    Returns (folio_DBFField, product_DBFField) or (None, None) if not found.
    """
    field_map = {f.name.upper(): f for f in fields}
    for folio_name, product_name in FOLIO_PRODUCT_CANDIDATES:
        if folio_name.upper() in field_map and product_name.upper() in field_map:
            return field_map[folio_name.upper()], field_map[product_name.upper()]
    return None, None

def get_csv_column_indices(headers: list):
    """
    Auto-detect PAN, folio, product column indices from CSV headers.
    Returns (pan_idx, folio_idx, product_idx) — any can be None if not found.
    """
    header_map = {h.strip().upper(): i for i, h in enumerate(headers)}

    # PAN — check candidates in order
    pan_idx = None
    for candidate in CSV_PAN_CANDIDATES:
        if candidate.upper() in header_map:
            pan_idx = header_map[candidate.upper()]
            break
    # Also try any column containing 'PAN'
    if pan_idx is None:
        for i, h in enumerate(headers):
            if 'PAN' in h.strip().upper():
                pan_idx = i
                break

    # FOLIO+PRODUCT — check candidates in order
    folio_idx = product_idx = None
    for folio_name, product_name in FOLIO_PRODUCT_CANDIDATES:
        fi = header_map.get(folio_name.upper())
        pi = header_map.get(product_name.upper())
        if fi is not None and pi is not None:
            folio_idx, product_idx = fi, pi
            break

    return pan_idx, folio_idx, product_idx


# ════════════════════════════════════════════════════════════════════════════
#  MASTER FILE READER
# ════════════════════════════════════════════════════════════════════════════

def _find_pan_org_in_rows(rows: list, sheet_name: str):
    """
    Scan up to 20 rows to find a row that has both a PAN and ORG column.
    Returns (pan_idx, org_idx, header_row_num) or (None, None, None).
    """
    for row_num, row in enumerate(rows[:20]):
        headers = [str(c).strip() if c is not None else "" for c in row]
        pi = oi = None
        for i, h in enumerate(headers):
            hu = h.upper()
            if pi is None and "PAN" in hu:
                pi = i
            if oi is None and "ORG" in hu:
                oi = i
        if pi is not None and oi is not None:
            log.info(f"  Sheet '{sheet_name}' — header row {row_num + 1}: {headers}")
            log.info(f"  PAN column -> '{headers[pi]}' | ORG column -> '{headers[oi]}'")
            return pi, oi, row_num
    return None, None, None


def load_pan_org_map(master_path: str, sheet_name: str = None) -> dict:
    """
    Read master XLSX. Auto-searches all sheets (or a specific sheet if given)
    for columns containing 'PAN' and 'ORG' in the name (case-insensitive).
    Returns { PAN_VALUE_UPPER: ORG_NAME }
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    log.info(f"Loading master file: {master_path}")
    t0 = time.perf_counter()

    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    all_sheet_names = wb.sheetnames
    log.info(f"  Sheets in workbook: {all_sheet_names}")

    # Decide which sheets to search
    if sheet_name:
        if sheet_name not in all_sheet_names:
            log.error(f"Sheet '{sheet_name}' not found. Available: {all_sheet_names}")
            wb.close()
            sys.exit(1)
        sheets_to_try = [sheet_name]
    else:
        sheets_to_try = all_sheet_names   # auto-search every sheet

    pan_idx = org_idx = header_row_num = None
    all_rows = []
    found_sheet = None

    for sname in sheets_to_try:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        pi, oi, hrn = _find_pan_org_in_rows(rows, sname)
        if pi is not None:
            pan_idx, org_idx, header_row_num = pi, oi, hrn
            all_rows = rows
            found_sheet = sname
            break

    wb.close()

    if pan_idx is None:
        log.error(
            f"Could not find PAN+ORG columns in any sheet.\n"
            f"  Sheets searched: {sheets_to_try}\n"
            f"  Column names must contain 'PAN' and 'ORG' (case-insensitive)."
        )
        sys.exit(1)

    log.info(f"  Using sheet: '{found_sheet}'")

    # Build map — rules:
    #   1. NA / #N/A / blank org values are ignored
    #   2. First real org found for a PAN wins (no overwrite)
    NA_VALUES = {"NA", "N/A", "#N/A", "#NA", "NONE", "NULL", "-", ""}
    pan_org = {}
    na_skipped = 0
    conflict_skipped = 0
    for row in all_rows[header_row_num + 1:]:
        if not any(row):
            continue
        pan_val = row[pan_idx] if pan_idx < len(row) else None
        org_val = row[org_idx] if org_idx < len(row) else None
        if pan_val is None:
            continue
        pan_str = str(pan_val).strip().upper()
        org_str = str(org_val).strip() if org_val is not None else ""
        org_str = org_str.title()   # normalise case: "scripbox"/"SCRIPBOX"/"Scripbox" -> "Scripbox"
        if not pan_str:
            continue
        # Skip NA/blank org values
        if org_str.upper() in NA_VALUES:
            na_skipped += 1
            continue
        # First real org wins — don't overwrite existing entry
        if pan_str in pan_org:
            if pan_org[pan_str] != org_str:
                conflict_skipped += 1
            continue
        pan_org[pan_str] = org_str
    if na_skipped:
        log.info(f"  Skipped {na_skipped:,} NA/blank org entries")
    if conflict_skipped:
        log.info(f"  Skipped {conflict_skipped:,} duplicate PAN entries (first org kept)")

    # ── Build FOLIO_PRODUCT→ORG fallback map ─────────────────────────────────
    # Looks for a column containing 'FOLIO_PRODUCT' in the master sheet
    folio_product_org = {}
    fp_col_idx = None
    if header_row_num is not None:
        header_row = all_rows[header_row_num]
        for i, cell in enumerate(header_row):
            if cell is not None and "FOLIO_PRODUCT" in str(cell).upper():
                fp_col_idx = i
                log.info(f"  Fallback column -> '{cell}' (index {i})")
                break

    if fp_col_idx is not None:
        fp_skipped = 0
        for row in all_rows[header_row_num + 1:]:
            if not any(row):
                continue
            fp_val  = row[fp_col_idx] if fp_col_idx < len(row) else None
            org_val = row[org_idx]    if org_idx    < len(row) else None
            if fp_val is None:
                continue
            fp_str  = str(fp_val).strip()
            org_str = str(org_val).strip() if org_val is not None else ""
            org_str = org_str.title()
            if not fp_str or org_str.upper() in NA_VALUES:
                fp_skipped += 1
                continue
            if fp_str not in folio_product_org:
                folio_product_org[fp_str] = org_str
        log.info(f"  Loaded {len(folio_product_org):,} FOLIO_PRODUCT->ORG fallback mappings")
    else:
        log.warning("  No FOLIO_PRODUCT column found in master — blank PAN rows will go to UNMATCHED")

    elapsed = time.perf_counter() - t0
    log.info(f"  Loaded {len(pan_org):,} PAN->ORG mappings in {elapsed:.2f}s")
    orgs = sorted(set(pan_org.values()))
    log.info(f"  Organisations ({len(orgs)}): {', '.join(orgs)}")
    return pan_org, folio_product_org


# ════════════════════════════════════════════════════════════════════════════
#  LOW-LEVEL DBF PARSER  (pure struct, zero-copy memoryview)
# ════════════════════════════════════════════════════════════════════════════

FIELD_ENTRY_SIZE = 32

class DBFField:
    __slots__ = ("name", "type", "length", "decimal", "offset")
    def __init__(self, name, ftype, length, decimal, offset):
        self.name    = name
        self.type    = ftype
        self.length  = length
        self.decimal = decimal
        self.offset  = offset   # byte offset within record (after deletion flag)


def _read_dbf_header(fh):
    """Returns (fields, num_records, header_size, record_size)."""
    raw = fh.read(32)
    num_records  = struct.unpack_from("<I", raw, 4)[0]
    header_size  = struct.unpack_from("<H", raw, 8)[0]
    record_size  = struct.unpack_from("<H", raw, 10)[0]

    fields = []
    field_offset = 0   # offset within field data (deletion flag handled separately)
    while True:
        entry = fh.read(FIELD_ENTRY_SIZE)
        if not entry or entry[0] == 0x0D:
            break
        if len(entry) < FIELD_ENTRY_SIZE:
            break
        name  = entry[0:11].split(b'\x00')[0].decode('ascii', errors='replace').strip()
        ftype = chr(entry[11])
        length  = entry[16]
        decimal = entry[17]
        fields.append(DBFField(name, ftype, length, decimal, field_offset))
        field_offset += length

    fh.seek(header_size)
    return fields, num_records, header_size, record_size


def _find_pan_field(fields):
    """Return first field whose name contains 'PAN' (case-insensitive)."""
    for f in fields:
        if "PAN" in f.name.upper():
            return f
    return None


# ════════════════════════════════════════════════════════════════════════════
#  DBF WRITER  (buffered, patches record count on close)
# ════════════════════════════════════════════════════════════════════════════

WRITE_BUFFER_ROWS = 4096

class DBFWriter:
    __slots__ = ("_fh", "_fields", "_record_size", "_count", "_buf", "_buf_n")

    def __init__(self, path: str, fields):
        self._fields      = fields
        self._record_size = 1 + sum(f.length for f in fields)
        self._count       = 0
        self._buf         = []
        self._buf_n       = 0
        self._fh          = open(path, "wb")
        self._write_header()

    def _write_header(self):
        today = date.today()
        num_fields  = len(self._fields)
        header_size = 32 + num_fields * FIELD_ENTRY_SIZE + 1
        self._fh.write(struct.pack(
            "<BBBBIHH20x",
            3,
            today.year - 1900, today.month, today.day,
            0,                  # placeholder record count
            header_size,
            self._record_size,
        ))
        for f in self._fields:
            name_b = f.name.encode('ascii', errors='replace').ljust(11, b'\x00')[:11]
            self._fh.write(
                name_b +
                f.type.encode('ascii') +
                b'\x00\x00\x00\x00' +
                bytes([f.length, f.decimal]) +
                b'\x00' * 14
            )
        self._fh.write(b'\x0D')

    def write_record(self, record_bytes: bytes):
        self._buf.append(b' ' + record_bytes)
        self._buf_n += 1
        self._count += 1
        if self._buf_n >= WRITE_BUFFER_ROWS:
            self._flush()

    def _flush(self):
        if self._buf:
            self._fh.write(b''.join(self._buf))
            self._buf.clear()
            self._buf_n = 0

    def close(self):
        self._flush()
        self._fh.write(b'\x1A')
        self._fh.seek(4)
        self._fh.write(struct.pack("<I", self._count))
        self._fh.close()


# ════════════════════════════════════════════════════════════════════════════
#  CORE PROCESSING  (one DBF file)
# ════════════════════════════════════════════════════════════════════════════

DATA_THRESHOLD = 500 * 1024 * 1024   # 500 MB — below this use bulk memoryview
READ_CHUNK_ROWS = 65536


def _safe_name(org: str) -> str:
    safe = "".join(c if (c.isalnum() or c in "-_") else "_" for c in org)
    return safe[:60]


def process_dbf(dbf_path: str, pan_org: dict, output_dir: str, folio_product_org: dict = None) -> dict:
    dbf_path   = Path(dbf_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name  = dbf_path.stem

    t0 = time.perf_counter()
    summary = {
        "file": dbf_path.name, "total": 0,
        "matched": 0, "unmatched": 0,
        "orgs": {}, "outputs": [], "error": None,
    }

    try:
        with open(dbf_path, "rb") as fh:
            fields, num_records, header_size, record_size = _read_dbf_header(fh)

            pan_field = _find_pan_field(fields)
            if pan_field is None:
                summary["error"] = f"No PAN field found. Fields: {[f.name for f in fields]}"
                return summary

            log.info(
                f"[{dbf_path.name}] {num_records:,} records | "
                f"PAN field: '{pan_field.name}' | record size: {record_size}B"
            )

            pan_start = pan_field.offset
            pan_end   = pan_start + pan_field.length

            # ── Auto-detect folio/product fields for blank PAN fallback ─────────
            folio_field, product_field = get_folio_product_fields(fields)
            if folio_field and product_field:
                log.info(f"  Fallback fields auto-detected: FOLIO='{folio_field.name}' PRODUCT='{product_field.name}'")
            else:
                log.warning(f"  No matching folio+product fields found in {dbf_path.name} — blank PAN rows -> UNMATCHED")

            writers = {}
            unmatched_writer = [None]
            blank_pan_fallback = [None]  # separate file for blank PAN rows that matched via fallback

            def get_writer(org_name):
                if org_name not in writers:
                    out = output_dir / f"{base_name}_{_safe_name(org_name)}.dbf"
                    writers[org_name] = DBFWriter(str(out), fields)
                    summary["outputs"].append(str(out))
                return writers[org_name]

            def get_unmatched():
                if unmatched_writer[0] is None:
                    out = output_dir / f"{base_name}_UNMATCHED.dbf"
                    unmatched_writer[0] = DBFWriter(str(out), fields)
                    summary["outputs"].append(str(out))
                return unmatched_writer[0]

            def handle_record(rec_mv):
                if rec_mv[0] == 0x2A:   # deleted
                    return
                pan_raw = bytes(rec_mv[1 + pan_start : 1 + pan_end])
                pan_str = pan_raw.decode('ascii', errors='replace').strip().upper()
                record_data = bytes(rec_mv[1:])
                summary["total"] += 1

                org = None

                # ── Step 1: Try FOLIO+PRODUCT first ──────────────────────────
                if folio_field and product_field and folio_product_org:
                    folio_raw   = bytes(rec_mv[1 + folio_field.offset   : 1 + folio_field.offset   + folio_field.length])
                    product_raw = bytes(rec_mv[1 + product_field.offset : 1 + product_field.offset + product_field.length])
                    folio_str   = folio_raw.decode('ascii', errors='replace').strip()
                    product_str = product_raw.decode('ascii', errors='replace').strip()
                    fp_key      = folio_str + product_str
                    org = folio_product_org.get(fp_key)
                    if org:
                        summary["fp_matched"] = summary.get("fp_matched", 0) + 1

                # ── Step 2: Fallback to PAN if FOLIO+PRODUCT didn't match ────
                if org is None and pan_str:
                    org = pan_org.get(pan_str)
                    if org:
                        summary["pan_matched"] = summary.get("pan_matched", 0) + 1

                # ── Step 3: Write result ──────────────────────────────────────
                if org:
                    get_writer(org).write_record(record_data)
                    summary["matched"] += 1
                    summary["orgs"][org] = summary["orgs"].get(org, 0) + 1
                else:
                    get_unmatched().write_record(record_data)
                    summary["unmatched"] += 1

            file_size = dbf_path.stat().st_size

            if file_size <= DATA_THRESHOLD:
                # Fast path: bulk read into memoryview
                raw = fh.read()
                mv  = memoryview(raw)
                n   = len(raw) // record_size
                for i in range(n):
                    handle_record(mv[i * record_size : (i + 1) * record_size])
            else:
                # Chunked path: large files
                chunk_bytes = record_size * READ_CHUNK_ROWS
                leftover = b""
                while True:
                    raw = fh.read(chunk_bytes)
                    if not raw:
                        break
                    raw = leftover + raw
                    n_complete = len(raw) // record_size
                    leftover   = raw[n_complete * record_size:]
                    mv = memoryview(raw)
                    for i in range(n_complete):
                        handle_record(mv[i * record_size : (i + 1) * record_size])

            for w in writers.values():
                w.close()
            if unmatched_writer[0]:
                unmatched_writer[0].close()

    except Exception as exc:
        summary["error"] = str(exc)
        log.error(f"[{dbf_path.name}] ERROR: {exc}")
        return summary

    elapsed = time.perf_counter() - t0
    rate = summary["total"] / elapsed if elapsed > 0 else 0
    log.info(
        f"[{dbf_path.name}] Done {elapsed:.2f}s | "
        f"{summary['total']:,} records @ {rate:,.0f} rec/s | "
        f"matched={summary['matched']:,} unmatched={summary['unmatched']:,}"
    )
    for org, cnt in sorted(summary["orgs"].items()):
        log.info(f"  -> {org}: {cnt:,} rows")
    return summary


# ════════════════════════════════════════════════════════════════════════════
#  CSV PROCESSING
# ════════════════════════════════════════════════════════════════════════════

def process_csv(csv_path: str, pan_org: dict, output_dir: str, folio_product_org: dict = None) -> dict:
    import csv as csv_module
    csv_path   = Path(csv_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name  = csv_path.stem

    t0 = time.perf_counter()
    summary = {
        "file": csv_path.name, "total": 0,
        "matched": 0, "unmatched": 0,
        "orgs": {}, "outputs": [], "error": None,
    }

    try:
        # Writers keyed by org name — write CSV per org
        writers  = {}   # org_name -> file handle + csv.writer
        unmatched_writer = [None]

        def get_writer(org_name):
            if org_name not in writers:
                out = output_dir / f"{base_name}_{_safe_name(org_name)}.csv"
                fh  = open(str(out), "w", newline="", encoding="utf-8-sig")
                w   = csv_module.writer(fh)
                writers[org_name] = (fh, w)
                summary["outputs"].append(str(out))
            return writers[org_name][1]

        def get_unmatched():
            if unmatched_writer[0] is None:
                out = output_dir / f"{base_name}_UNMATCHED.csv"
                fh  = open(str(out), "w", newline="", encoding="utf-8-sig")
                w   = csv_module.writer(fh)
                unmatched_writer[0] = (fh, w)
                summary["outputs"].append(str(out))
            return unmatched_writer[0][1]

        NA_VALUES = {"NA", "N/A", "#N/A", "#NA", "NONE", "NULL", "-", ""}

        with open(str(csv_path), "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv_module.reader(f)
            headers = next(reader)

            # Auto-detect columns
            pan_idx, folio_idx, product_idx = get_csv_column_indices(headers)

            if pan_idx is None:
                summary["error"] = f"No PAN column found. Headers: {headers}"
                return summary

            log.info(
                f"[{csv_path.name}] PAN col='{headers[pan_idx]}' | "
                f"FOLIO col='{headers[folio_idx] if folio_idx is not None else 'N/A'}' | "
                f"PRODUCT col='{headers[product_idx] if product_idx is not None else 'N/A'}'"
            )

            # Write header row to all output files lazily (written on first data row)
            header_written = {}

            def ensure_header(org_name, writer):
                if org_name not in header_written:
                    writer.writerow(headers)
                    header_written[org_name] = True

            def ensure_unmatched_header(writer):
                if "_unmatched" not in header_written:
                    writer.writerow(headers)
                    header_written["_unmatched"] = True

            for row in reader:
                if not any(r.strip() for r in row):
                    continue  # skip empty rows
                summary["total"] += 1

                pan_str  = row[pan_idx].strip().upper() if pan_idx < len(row) else ""
                org      = None

                # Step 1: FOLIO+PRODUCT lookup
                if folio_idx is not None and product_idx is not None and folio_product_org:
                    folio_str   = row[folio_idx].strip()   if folio_idx   < len(row) else ""
                    product_str = row[product_idx].strip() if product_idx < len(row) else ""
                    fp_key      = folio_str + product_str
                    org = folio_product_org.get(fp_key)
                    if org:
                        summary["fp_matched"] = summary.get("fp_matched", 0) + 1

                # Step 2: PAN fallback
                if org is None and pan_str and pan_str.upper() not in NA_VALUES:
                    org = pan_org.get(pan_str)
                    if org:
                        summary["pan_matched"] = summary.get("pan_matched", 0) + 1

                # Step 3: Write result
                if org:
                    w = get_writer(org)
                    ensure_header(org, w)
                    w.writerow(row)
                    summary["matched"] += 1
                    summary["orgs"][org] = summary["orgs"].get(org, 0) + 1
                else:
                    w = get_unmatched()
                    ensure_unmatched_header(w)
                    w.writerow(row)
                    summary["unmatched"] += 1

        # Close all writers
        for fh, _ in writers.values():
            fh.close()
        if unmatched_writer[0]:
            unmatched_writer[0][0].close()

    except Exception as exc:
        summary["error"] = str(exc)
        log.error(f"[{csv_path.name}] ERROR: {exc}")
        return summary

    elapsed = time.perf_counter() - t0
    rate = summary["total"] / elapsed if elapsed > 0 else 0
    log.info(
        f"[{csv_path.name}] Done {elapsed:.2f}s | "
        f"{summary['total']:,} records @ {rate:,.0f} rec/s | "
        f"matched={summary['matched']:,} unmatched={summary['unmatched']:,}"
    )
    for org, cnt in sorted(summary["orgs"].items()):
        log.info(f"  -> {org}: {cnt:,} rows")
    return summary


# ════════════════════════════════════════════════════════════════════════════
#  PARALLEL RUNNER
# ════════════════════════════════════════════════════════════════════════════

def run(master_path, dbf_paths, csv_paths, output_dir, workers, sheet_name=None):
    t_total = time.perf_counter()
    pan_org, folio_product_org = load_pan_org_map(master_path, sheet_name)

    if not pan_org and not folio_product_org:
        log.error("Master map is empty — nothing to do.")
        sys.exit(1)

    all_tasks = (
        [(p, "dbf") for p in dbf_paths] +
        [(p, "csv") for p in csv_paths]
    )
    log.info(f"\nProcessing {len(dbf_paths)} DBF + {len(csv_paths)} CSV file(s) with {workers} worker(s)...")

    summaries = []
    if workers == 1 or len(all_tasks) == 1:
        for p, ftype in all_tasks:
            if ftype == "dbf":
                summaries.append(process_dbf(p, pan_org, output_dir, folio_product_org))
            else:
                summaries.append(process_csv(p, pan_org, output_dir, folio_product_org))
    else:
        with ProcessPoolExecutor(max_workers=workers) as pool:
            futures = {}
            for p, ftype in all_tasks:
                if ftype == "dbf":
                    futures[pool.submit(process_dbf, p, pan_org, output_dir, folio_product_org)] = p
                else:
                    futures[pool.submit(process_csv, p, pan_org, output_dir, folio_product_org)] = p
            for fut in as_completed(futures):
                summaries.append(fut.result())

    total_elapsed   = time.perf_counter() - t_total
    total_records   = sum(s["total"]   for s in summaries)
    total_matched   = sum(s["matched"] for s in summaries)
    total_unmatched = sum(s["unmatched"] for s in summaries)
    total_outputs   = sum(len(s["outputs"]) for s in summaries)
    errors = [s for s in summaries if s["error"]]

    print("\n" + "=" * 60)
    print("  SEGREGATION COMPLETE")
    print("=" * 60)
    total_fp_matched  = sum(s.get("fp_matched",  0) for s in summaries)
    total_pan_matched = sum(s.get("pan_matched", 0) for s in summaries)
    print(f"  Total records processed : {total_records:>12,}")
    print(f"  Matched rows            : {total_matched:>12,}")
    print(f"    via FOLIO+PRODUCT     : {total_fp_matched:>12,}")
    print(f"    via PAN (fallback)    : {total_pan_matched:>12,}")
    print(f"  Unmatched rows          : {total_unmatched:>12,}")
    print(f"  Output files created    : {total_outputs:>12,}")
    print(f"  Time elapsed            : {total_elapsed:>11.2f}s")
    if errors:
        print(f"\n  ERRORS ({len(errors)}):")
        for s in errors:
            print(f"    {s['file']}: {s['error']}")
    print(f"\n  Output folder: {Path(output_dir).resolve()}")
    print("=" * 60)


# ════════════════════════════════════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Segregate DBF files by organisation via master XLSX PAN->ORG map.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  py segregate.py --master May22.xlsx --dbf file1.dbf
  py segregate.py --master May22.xlsx --dbf file1.dbf file2.dbf file3.dbf
  py segregate.py --master May22.xlsx --dbf file1.dbf --sheet May22
  py segregate.py --master May22.xlsx --dbf file1.dbf --output C:\\results
        """
    )
    parser.add_argument("--master",  "-m", required=True, help="Master XLSX file path")
    parser.add_argument("--dbf",     "-d", nargs="+", default=[], help="DBF file(s)")
    parser.add_argument("--csv",     "-c", nargs="+", default=[], help="CSV file(s)")
    parser.add_argument("--output",  "-o", default="./output", help="Output folder (default: ./output)")
    parser.add_argument("--sheet",   "-s", default=None, help="Specific sheet name in master XLSX (optional — auto-searches all sheets if omitted)")
    parser.add_argument("--workers", "-w", type=int, default=None, help="Parallel workers (default: min(total files, 4))")
    parser.add_argument("--diagnose", action="store_true", help="Print sample PAN values from master and first DBF to debug mismatches")
    args = parser.parse_args()

    if not args.dbf and not args.csv:
        log.error("Provide at least one --dbf or --csv file.")
        sys.exit(1)

    if not Path(args.master).exists():
        log.error(f"Master file not found: {args.master}")
        sys.exit(1)
    missing = [p for p in args.dbf + args.csv if not Path(p).exists()]
    if missing:
        log.error(f"File(s) not found: {', '.join(missing)}")
        sys.exit(1)

    if args.diagnose:
        target = args.dbf[0] if args.dbf else args.csv[0]
        diagnose(args.master, target, args.sheet)
        return

    workers = args.workers or min(len(args.dbf) + len(args.csv), 4)
    run(args.master, args.dbf, args.csv, args.output, workers, args.sheet)


if __name__ == "__main__":
    main()


# ════════════════════════════════════════════════════════════════════════════
#  DIAGNOSE MODE  — prints sample PAN values from master + DBF for comparison
# ════════════════════════════════════════════════════════════════════════════

def diagnose(master_path: str, dbf_path: str, sheet_name: str = None):
    import openpyxl

    print("\n" + "=" * 60)
    print("  DIAGNOSE MODE")
    print("=" * 60)

    # ── Master samples ───────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    all_sheet_names = wb.sheetnames
    sheets_to_try = [sheet_name] if sheet_name else all_sheet_names

    pan_idx = org_idx = header_row_num = None
    all_rows = []
    found_sheet = None
    for sname in sheets_to_try:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        pi, oi, hrn = _find_pan_org_in_rows(rows, sname)
        if pi is not None:
            pan_idx, org_idx, header_row_num = pi, oi, hrn
            all_rows = rows
            found_sheet = sname
            break
    wb.close()

    print(f"\n[MASTER] Sheet: '{found_sheet}' | PAN col idx={pan_idx} | ORG col idx={org_idx}")
    print(f"  First 10 PAN values from master:")
    count = 0
    for row in all_rows[header_row_num + 1:]:
        if not any(row): continue
        pan_val = row[pan_idx] if pan_idx < len(row) else None
        org_val = row[org_idx] if org_idx < len(row) else None
        if pan_val is None: continue
        pan_str = str(pan_val).strip()
        org_str = str(org_val).strip() if org_val else ""
        print(f"    repr={repr(pan_str):<25}  len={len(pan_str):<4}  org={org_str}")
        count += 1
        if count >= 10: break

    # ── DBF samples ──────────────────────────────────────────────────────────
    print(f"\n[DBF] File: {dbf_path}")
    with open(dbf_path, "rb") as fh:
        fields, num_records, header_size, record_size = _read_dbf_header(fh)
        pan_field = _find_pan_field(fields)
        if pan_field is None:
            print(f"  ERROR: No PAN field. Fields: {[f.name for f in fields]}")
            return
        print(f"  PAN field: '{pan_field.name}' | length={pan_field.length} | offset={pan_field.offset}")
        print(f"  All field names: {[f.name for f in fields]}")
        print(f"  First 10 PAN values from DBF:")
        raw = fh.read()
        mv  = memoryview(raw)
        n   = len(raw) // record_size
        count = 0
        for i in range(n):
            rec = mv[i * record_size : (i + 1) * record_size]
            if rec[0] == 0x2A: continue
            pan_raw = bytes(rec[1 + pan_field.offset : 1 + pan_field.offset + pan_field.length])
            pan_str = pan_raw.decode('ascii', errors='replace').strip()
            print(f"    repr={repr(pan_str):<25}  len={len(pan_str)}")
            count += 1
            if count >= 10: break

    print("\n" + "=" * 60)
    print("  Compare the repr() values above — spaces, dots, or length")
    print("  differences between master and DBF will cause mismatches.")
    print("=" * 60 + "\n")